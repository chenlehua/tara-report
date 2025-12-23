"""
TARA Excel报告生成器
用于生成车载信息娱乐系统(IVI)的威胁分析和风险评估报告
"""
import os
from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image


# ==================== 样式定义 ====================
class TARAStyles:
    """TARA报告样式常量"""
    # 颜色定义
    DARK_BLUE = "FF2F5496"
    MEDIUM_BLUE = "FF4472C4"
    LIGHT_BLUE = "FF8EA9DB"
    WHITE = "FFFFFFFF"
    
    # 字体
    TITLE_FONT = Font(name='等线', size=16, bold=True, color=DARK_BLUE)
    SECTION_FONT = Font(name='等线', size=10, bold=True, color=WHITE)
    HEADER_FONT = Font(name='等线', size=10, bold=True, color=WHITE)
    SUBHEADER_FONT = Font(name='等线', size=10, bold=True, color=WHITE)
    NORMAL_FONT = Font(name='等线', size=11)
    
    # 填充
    SECTION_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')
    HEADER_FILL = PatternFill(start_color=MEDIUM_BLUE, end_color=MEDIUM_BLUE, fill_type='solid')
    SUBHEADER_FILL = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type='solid')
    
    # 边框
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 对齐
    CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)
    TOP_LEFT_ALIGN = Alignment(horizontal='left', vertical='top', wrap_text=True)


# ==================== Sheet 0: 封面 ====================
def create_cover_sheet(wb: Workbook, data: Dict[str, Any]) -> None:
    """创建封面Sheet"""
    ws = wb.active
    ws.title = "0. 封面 Front Cover"
    
    # 设置列宽
    col_widths = [15, 15, 15, 15, 8, 25, 8]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # 数据等级信息
    ws['F4'] = f"数据等级：{data.get('data_level', '秘密')}\nData level: Confidential"
    ws['F4'].font = Font(bold=True)
    ws['F4'].alignment = TARAStyles.LEFT_ALIGN
    
    ws['F5'] = f"编号：{data.get('document_number', '')}\nNumber: {data.get('document_number', '')}"
    ws['F5'].font = Font(bold=True)
    ws['F5'].alignment = TARAStyles.LEFT_ALIGN
    
    ws['F6'] = f"版本：{data.get('version', '')}\nVersion："
    ws['F6'].font = Font(bold=True)
    ws['F6'].alignment = TARAStyles.LEFT_ALIGN
    
    # 主标题
    ws.merge_cells('A7:G7')
    ws['A7'] = f"{data.get('report_title', '威胁分析和风险评估报告')}\n{data.get('report_title_en', 'Threat Analysis And Risk Assessment Report')}"
    ws['A7'].font = Font(size=16, bold=True)
    ws['A7'].alignment = TARAStyles.CENTER_ALIGN
    ws.row_dimensions[7].height = 45
    
    # 项目名称
    ws.merge_cells('E8:G8')
    ws['E8'] = data.get('project_name', '')
    ws['E8'].font = Font(size=16, bold=True)
    ws['E8'].alignment = TARAStyles.CENTER_ALIGN
    
    # 签名信息
    sign_info = [
        ('A9:B9', 'C9:D9', '编制/日期：\nAuthor/Date', data.get('author_date', '')),
        ('A10:B10', 'C10:D10', '审核/日期：\nReview/Date', data.get('review_date', '')),
        ('A11:B11', 'C11:D11', '会签/日期：\nSignature/Date', data.get('sign_date', '')),
        ('A12:B12', 'C12:D12', '批准/日期：\nApprove/Date', data.get('approve_date', '')),
    ]
    
    for label_range, value_range, label, value in sign_info:
        ws.merge_cells(label_range)
        ws.merge_cells(value_range)
        start_cell = label_range.split(':')[0]
        value_cell = value_range.split(':')[0]
        ws[start_cell] = label
        ws[start_cell].alignment = TARAStyles.LEFT_ALIGN
        ws[value_cell] = value
        ws[value_cell].alignment = TARAStyles.LEFT_ALIGN


# ==================== Sheet 1: 相关定义 ====================
def create_definitions_sheet(wb: Workbook, data: Dict[str, Any]) -> None:
    """创建相关定义Sheet"""
    ws = wb.create_sheet("1-相关定义")
    
    # 设置列宽
    col_widths = [15, 20, 20, 20, 20, 30]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    current_row = 1
    
    # 标题
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'] = data.get('title', 'MY25 EV平台中控主机 TARA分析报告 - 相关定义')
    ws[f'A{current_row}'].font = TARAStyles.TITLE_FONT
    ws[f'A{current_row}'].alignment = TARAStyles.CENTER_ALIGN
    ws.row_dimensions[current_row].height = 30
    current_row += 2
    
    # 1. 功能描述
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'] = "1. 功能描述 Functional Description"
    ws[f'A{current_row}'].font = TARAStyles.SECTION_FONT
    ws[f'A{current_row}'].fill = TARAStyles.SECTION_FILL
    current_row += 1
    
    # 功能描述内容
    ws.merge_cells(f'A{current_row}:F{current_row + 6}')
    ws[f'A{current_row}'] = data.get('functional_description', '')
    ws[f'A{current_row}'].font = Font(size=11)
    ws[f'A{current_row}'].alignment = TARAStyles.TOP_LEFT_ALIGN
    current_row += 8
    
    # 2. 项目边界
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'] = "2. 项目边界 Item Boundary"
    ws[f'A{current_row}'].font = TARAStyles.SECTION_FONT
    ws[f'A{current_row}'].fill = TARAStyles.SECTION_FILL
    current_row += 1
    
    # 项目边界图片区域
    ws.merge_cells(f'A{current_row}:F{current_row + 17}')
    if data.get('item_boundary_image') and os.path.exists(data['item_boundary_image']):
        try:
            img = Image(data['item_boundary_image'])
            img.width = 700
            img.height = 350
            ws.add_image(img, f'A{current_row}')
        except Exception:
            ws[f'A{current_row}'] = f"[图片: {data['item_boundary_image']}]"
    current_row += 19
    
    # 3. 系统架构图
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'] = "3. 系统架构图 System Architecture"
    ws[f'A{current_row}'].font = TARAStyles.SECTION_FONT
    ws[f'A{current_row}'].fill = TARAStyles.SECTION_FILL
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:F{current_row + 17}')
    if data.get('system_architecture_image') and os.path.exists(data['system_architecture_image']):
        try:
            img = Image(data['system_architecture_image'])
            img.width = 700
            img.height = 350
            ws.add_image(img, f'A{current_row}')
        except Exception:
            ws[f'A{current_row}'] = f"[图片: {data['system_architecture_image']}]"
    current_row += 19
    
    # 4. 软件架构图
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'] = "4. 软件架构图 Software Architecture"
    ws[f'A{current_row}'].font = TARAStyles.SECTION_FONT
    ws[f'A{current_row}'].fill = TARAStyles.SECTION_FILL
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:F{current_row + 17}')
    if data.get('software_architecture_image') and os.path.exists(data['software_architecture_image']):
        try:
            img = Image(data['software_architecture_image'])
            img.width = 700
            img.height = 350
            ws.add_image(img, f'A{current_row}')
        except Exception:
            ws[f'A{current_row}'] = f"[图片: {data['software_architecture_image']}]"
    current_row += 19
    
    # 5. 相关项假设
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'] = "5. 相关项假设 Item Assumptions"
    ws[f'A{current_row}'].font = TARAStyles.SECTION_FONT
    ws[f'A{current_row}'].fill = TARAStyles.SECTION_FILL
    current_row += 1
    
    # 假设表头
    ws[f'A{current_row}'] = "假设编号\nAssumption ID"
    ws[f'A{current_row}'].font = TARAStyles.HEADER_FONT
    ws[f'A{current_row}'].fill = TARAStyles.HEADER_FILL
    ws[f'A{current_row}'].alignment = TARAStyles.CENTER_ALIGN
    ws[f'A{current_row}'].border = TARAStyles.THIN_BORDER
    
    ws.merge_cells(f'B{current_row}:F{current_row}')
    ws[f'B{current_row}'] = "假设描述 Assumption Description"
    ws[f'B{current_row}'].font = TARAStyles.HEADER_FONT
    ws[f'B{current_row}'].fill = TARAStyles.HEADER_FILL
    ws[f'B{current_row}'].alignment = TARAStyles.CENTER_ALIGN
    ws[f'B{current_row}'].border = TARAStyles.THIN_BORDER
    current_row += 1
    
    # 假设数据
    for assumption in data.get('assumptions', []):
        ws[f'A{current_row}'] = assumption.get('id', '')
        ws[f'A{current_row}'].border = TARAStyles.THIN_BORDER
        ws[f'A{current_row}'].alignment = TARAStyles.CENTER_ALIGN
        
        ws.merge_cells(f'B{current_row}:F{current_row}')
        ws[f'B{current_row}'] = assumption.get('description', '')
        ws[f'B{current_row}'].border = TARAStyles.THIN_BORDER
        ws[f'B{current_row}'].alignment = TARAStyles.LEFT_ALIGN
        current_row += 1
    
    current_row += 1
    
    # 6. 术语表
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'] = "6. 术语表 Terminology"
    ws[f'A{current_row}'].font = TARAStyles.SECTION_FONT
    ws[f'A{current_row}'].fill = TARAStyles.SECTION_FILL
    current_row += 1
    
    # 术语表头
    ws[f'A{current_row}'] = "缩写\nAbbreviation"
    ws[f'A{current_row}'].font = TARAStyles.HEADER_FONT
    ws[f'A{current_row}'].fill = TARAStyles.HEADER_FILL
    ws[f'A{current_row}'].alignment = TARAStyles.CENTER_ALIGN
    ws[f'A{current_row}'].border = TARAStyles.THIN_BORDER
    
    ws.merge_cells(f'B{current_row}:E{current_row}')
    ws[f'B{current_row}'] = "英文全称 English Full Name"
    ws[f'B{current_row}'].font = TARAStyles.HEADER_FONT
    ws[f'B{current_row}'].fill = TARAStyles.HEADER_FILL
    ws[f'B{current_row}'].alignment = TARAStyles.CENTER_ALIGN
    ws[f'B{current_row}'].border = TARAStyles.THIN_BORDER
    
    ws[f'F{current_row}'] = "中文全称 Chinese Name"
    ws[f'F{current_row}'].font = TARAStyles.HEADER_FONT
    ws[f'F{current_row}'].fill = TARAStyles.HEADER_FILL
    ws[f'F{current_row}'].alignment = TARAStyles.CENTER_ALIGN
    ws[f'F{current_row}'].border = TARAStyles.THIN_BORDER
    current_row += 1
    
    # 术语数据
    for term in data.get('terminology', []):
        ws[f'A{current_row}'] = term.get('abbreviation', '')
        ws[f'A{current_row}'].border = TARAStyles.THIN_BORDER
        ws[f'A{current_row}'].alignment = TARAStyles.CENTER_ALIGN
        
        ws.merge_cells(f'B{current_row}:E{current_row}')
        ws[f'B{current_row}'] = term.get('english', '')
        ws[f'B{current_row}'].border = TARAStyles.THIN_BORDER
        ws[f'B{current_row}'].alignment = TARAStyles.LEFT_ALIGN
        
        ws[f'F{current_row}'] = term.get('chinese', '')
        ws[f'F{current_row}'].border = TARAStyles.THIN_BORDER
        ws[f'F{current_row}'].alignment = TARAStyles.LEFT_ALIGN
        current_row += 1


# ==================== Sheet 2: 资产列表&数据流图 ====================
def create_assets_sheet(wb: Workbook, data: Dict[str, Any]) -> None:
    """创建资产列表&数据流图Sheet"""
    ws = wb.create_sheet("2-资产列表&数据流图")
    
    # 设置列宽
    col_widths = [10, 15, 12, 50, 12, 12, 15, 12, 12, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    current_row = 1
    
    # 标题
    ws.merge_cells(f'A{current_row}:J{current_row}')
    ws[f'A{current_row}'] = data.get('title', 'MY25 EV平台中控主机- 资产列表 Asset List')
    ws[f'A{current_row}'].font = Font(size=14, bold=True, color=TARAStyles.DARK_BLUE)
    ws[f'A{current_row}'].alignment = TARAStyles.CENTER_ALIGN
    ws.row_dimensions[current_row].height = 25
    current_row += 2
    
    # 分组表头
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'] = "Asset Identification 资产识别"
    ws[f'A{current_row}'].font = TARAStyles.SECTION_FONT
    ws[f'A{current_row}'].fill = TARAStyles.SECTION_FILL
    ws[f'A{current_row}'].alignment = TARAStyles.CENTER_ALIGN
    ws[f'A{current_row}'].border = TARAStyles.THIN_BORDER
    
    ws.merge_cells(f'E{current_row}:J{current_row}')
    ws[f'E{current_row}'] = "Cybersecurity Attributes 网络安全属性"
    ws[f'E{current_row}'].font = TARAStyles.SECTION_FONT
    ws[f'E{current_row}'].fill = TARAStyles.SECTION_FILL
    ws[f'E{current_row}'].alignment = TARAStyles.CENTER_ALIGN
    ws[f'E{current_row}'].border = TARAStyles.THIN_BORDER
    current_row += 1
    
    # 详细表头
    headers = [
        ('A', '资产ID\nAsset ID'),
        ('B', '资产名称\nAsset Name'),
        ('C', '分类\nCategory'),
        ('D', '备注\nRemarks'),
        ('E', '真实性\nAuthenticity'),
        ('F', '完整性\nIntegrity'),
        ('G', '不可抵赖性\nNon-repudiation'),
        ('H', '机密性\nConfidentiality'),
        ('I', '可用性\nAvailability'),
        ('J', '权限\nAuthorization')
    ]
    
    for col, header in headers:
        ws[f'{col}{current_row}'] = header
        ws[f'{col}{current_row}'].font = TARAStyles.HEADER_FONT
        ws[f'{col}{current_row}'].fill = TARAStyles.HEADER_FILL
        ws[f'{col}{current_row}'].alignment = TARAStyles.CENTER_ALIGN
        ws[f'{col}{current_row}'].border = TARAStyles.THIN_BORDER
    ws.row_dimensions[current_row].height = 35
    current_row += 1
    
    # 资产数据
    for asset in data.get('assets', []):
        ws[f'A{current_row}'] = asset.get('id', '')
        ws[f'B{current_row}'] = asset.get('name', '')
        ws[f'C{current_row}'] = asset.get('category', '')
        ws[f'D{current_row}'] = asset.get('remarks', '')
        ws[f'E{current_row}'] = '√' if asset.get('authenticity') else ''
        ws[f'F{current_row}'] = '√' if asset.get('integrity') else ''
        ws[f'G{current_row}'] = '√' if asset.get('non_repudiation') else ''
        ws[f'H{current_row}'] = '√' if asset.get('confidentiality') else ''
        ws[f'I{current_row}'] = '√' if asset.get('availability') else ''
        ws[f'J{current_row}'] = '√' if asset.get('authorization') else ''
        
        for col in 'ABCDEFGHIJ':
            ws[f'{col}{current_row}'].border = TARAStyles.THIN_BORDER
            ws[f'{col}{current_row}'].alignment = TARAStyles.CENTER_ALIGN if col not in ['D'] else TARAStyles.LEFT_ALIGN
        current_row += 1
    
    # 数据流图
    current_row += 2
    if data.get('dataflow_image') and os.path.exists(data['dataflow_image']):
        try:
            img = Image(data['dataflow_image'])
            img.width = 800
            img.height = 400
            ws.add_image(img, f'A{current_row}')
        except Exception:
            ws[f'A{current_row}'] = f"[数据流图: {data['dataflow_image']}]"


# ==================== Sheet 3: 攻击树图 ====================
def create_attack_trees_sheet(wb: Workbook, data: Dict[str, Any]) -> None:
    """创建攻击树图Sheet"""
    ws = wb.create_sheet("3-攻击树图")
    
    # 设置列宽
    for i in range(1, 7):
        ws.column_dimensions[get_column_letter(i)].width = 20
    
    current_row = 1
    
    # 主标题
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'] = data.get('title', 'MY25 EV平台中控主机 - 攻击树分析 Attack Tree Analysis')
    ws[f'A{current_row}'].font = Font(size=14, bold=True, color=TARAStyles.DARK_BLUE)
    ws[f'A{current_row}'].alignment = TARAStyles.CENTER_ALIGN
    ws.row_dimensions[current_row].height = 25
    current_row += 2
    
    # 攻击树
    for tree in data.get('attack_trees', []):
        # 攻击树标题
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = tree.get('title', '')
        ws[f'A{current_row}'].font = TARAStyles.SECTION_FONT
        ws[f'A{current_row}'].fill = TARAStyles.SECTION_FILL
        ws[f'A{current_row}'].alignment = TARAStyles.CENTER_ALIGN
        current_row += 1
        
        # 攻击树图片区域
        ws.merge_cells(f'A{current_row}:F{current_row + 17}')
        if tree.get('image') and os.path.exists(tree['image']):
            try:
                img = Image(tree['image'])
                img.width = 700
                img.height = 350
                ws.add_image(img, f'A{current_row}')
            except Exception:
                ws[f'A{current_row}'] = f"[攻击树图: {tree['image']}]"
        current_row += 20


# ==================== Sheet 4: TARA分析结果 ====================
def create_tara_results_sheet(wb: Workbook, data: Dict[str, Any]) -> None:
    """创建TARA分析结果Sheet"""
    ws = wb.create_sheet("4-TARA分析结果")
    
    # 设置列宽
    col_widths = {
        'A': 8, 'B': 12, 'C': 10, 'D': 13, 'E': 13, 'F': 15,
        'G': 27, 'H': 12, 'I': 54, 'J': 82, 'K': 10,
        'L': 13, 'M': 8, 'N': 10, 'O': 8, 'P': 10, 'Q': 8,
        'R': 10, 'S': 8, 'T': 8, 'U': 10,
        'V': 14, 'W': 18, 'X': 6, 'Y': 14, 'Z': 28, 'AA': 6,
        'AB': 14, 'AC': 12, 'AD': 6, 'AE': 14, 'AF': 24, 'AG': 6,
        'AH': 8, 'AI': 10, 'AJ': 13, 'AK': 12, 'AL': 18, 'AM': 25, 'AN': 12
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    current_row = 1
    
    # 主标题
    ws.merge_cells(f'A{current_row}:AN{current_row}')
    ws[f'A{current_row}'] = data.get('title', 'MY25 EV平台中控主机_TARA分析结果 TARA Analysis Results')
    ws[f'A{current_row}'].font = Font(size=14, bold=True, color=TARAStyles.DARK_BLUE)
    ws[f'A{current_row}'].alignment = TARAStyles.CENTER_ALIGN
    current_row += 2
    
    # 第一层表头 (分组)
    group_headers = [
        ('A', 'F', 'Asset Identification资产识别'),
        ('G', 'K', 'Threat & Damage Scenario\n威胁&损害场景'),
        ('L', 'U', 'Threat Analysis\n威胁分析'),
        ('V', 'AI', 'Impact Analysis\n影响分析'),
        ('AJ', 'AJ', 'Risk Assessment\n风险评估'),
        ('AK', 'AK', 'Risk Treatment\n风险处置'),
        ('AL', 'AN', 'Risk Mitigation\n风险缓解')
    ]
    
    for start_col, end_col, title in group_headers:
        if start_col != end_col:
            ws.merge_cells(f'{start_col}{current_row}:{end_col}{current_row}')
        ws[f'{start_col}{current_row}'] = title
        ws[f'{start_col}{current_row}'].font = TARAStyles.SECTION_FONT
        ws[f'{start_col}{current_row}'].fill = TARAStyles.SECTION_FILL
        ws[f'{start_col}{current_row}'].alignment = TARAStyles.CENTER_ALIGN
        ws[f'{start_col}{current_row}'].border = TARAStyles.THIN_BORDER
    current_row += 1
    
    # 第二层表头
    header_row = current_row
    headers_row4 = [
        ('A', 'A', 'Asset\nID\n资产ID'),
        ('B', 'B', 'Asset Name\n资产名称'),
        ('C', 'E', '细分类'),
        ('F', 'F', 'Category\n分类'),
        ('G', 'G', 'Security Attributes\n安全属性'),
        ('H', 'H', 'STRIDE Model\nSTRIDE模型'),
        ('I', 'I', 'Potential Threat and Damage Scenario\n潜在威胁和损害场景'),
        ('J', 'J', 'Attack Path\n攻击路径'),
        ('K', 'K', '来源\n'),
        ('L', 'M', 'Attack Vector(V)\n攻击向量'),
        ('N', 'O', 'Attack Complexity(C)\n攻击复杂度'),
        ('P', 'Q', 'Privileges Required(P)\n权限要求'),
        ('R', 'S', 'User Interaction(U)\n用户交互'),
        ('T', 'U', 'Attack Feasibility\n攻击可行性计算'),
        ('V', 'X', 'Safety\n安全'),
        ('Y', 'AA', 'Financial\n经济'),
        ('AB', 'AD', 'Operational\n操作'),
        ('AE', 'AG', 'Privacy & Legislation\n隐私和法律'),
        ('AH', 'AI', 'Impact Level Calculation\n影响等级计算'),
        ('AJ', 'AJ', 'Risk Level\n风险等级'),
        ('AK', 'AK', 'Risk Treatment Decision\n风险处置决策'),
        ('AL', 'AL', 'Security Goal\n安全目标'),
        ('AM', 'AM', 'Security Requirement\n安全需求'),
        ('AN', 'AN', 'Source来源\n')
    ]
    
    for start_col, end_col, title in headers_row4:
        if start_col != end_col:
            ws.merge_cells(f'{start_col}{current_row}:{end_col}{current_row}')
        ws[f'{start_col}{current_row}'] = title
        ws[f'{start_col}{current_row}'].font = TARAStyles.HEADER_FONT
        ws[f'{start_col}{current_row}'].fill = TARAStyles.HEADER_FILL
        ws[f'{start_col}{current_row}'].alignment = TARAStyles.CENTER_ALIGN
        ws[f'{start_col}{current_row}'].border = TARAStyles.THIN_BORDER
    current_row += 1
    
    # 第三层表头 (子列)
    subheaders = [
        ('C', '子领域一'), ('D', '子领域二'), ('E', '子领域三'),
        ('K', 'WP29威胁映射'),
        ('L', 'Time\n内容'), ('M', 'Value\n指标值'),
        ('N', 'Content\n内容'), ('O', 'Value\n指标值'),
        ('P', 'Level\n等级'), ('Q', 'Value\n指标值'),
        ('R', 'Level\n等级'), ('S', 'Value\n指标值'),
        ('T', 'Calculation\n计算值'), ('U', 'Level\n等级'),
        ('V', 'Content\n内容'), ('W', 'Notes\n注释'), ('X', 'Value\n指标值'),
        ('Y', 'Content\n内容'), ('Z', 'Notes\n注释'), ('AA', 'Value\n指标值'),
        ('AB', 'Content\n内容'), ('AC', 'Notes\n注释'), ('AD', 'Value\n指标值'),
        ('AE', 'Content\n内容'), ('AF', 'Notes\n注释'), ('AG', 'Value\n指标值'),
        ('AH', 'Impact Calc\n影响计算'), ('AI', 'Impact Level\n影响等级'),
        ('AN', 'WP29 Control Mapping')
    ]
    
    for col, title in subheaders:
        ws[f'{col}{current_row}'] = title
        ws[f'{col}{current_row}'].font = TARAStyles.SUBHEADER_FONT
        ws[f'{col}{current_row}'].fill = TARAStyles.SUBHEADER_FILL
        ws[f'{col}{current_row}'].alignment = TARAStyles.CENTER_ALIGN
        ws[f'{col}{current_row}'].border = TARAStyles.THIN_BORDER
    
    # 合并跨行的表头单元格
    merge_headers = ['A', 'B', 'F', 'G', 'H', 'I', 'J', 'AJ', 'AK', 'AL', 'AM']
    for col in merge_headers:
        ws.merge_cells(f'{col}{header_row}:{col}{current_row}')
    
    current_row += 1
    
    # 数据行
    for result in data.get('results', []):
        row = current_row
        
        # 基本信息
        ws[f'A{row}'] = result.get('asset_id', '')
        ws[f'B{row}'] = result.get('asset_name', '')
        ws[f'C{row}'] = result.get('subdomain1', '')
        ws[f'D{row}'] = result.get('subdomain2', '')
        ws[f'E{row}'] = result.get('subdomain3', '')
        ws[f'F{row}'] = result.get('category', '')
        ws[f'G{row}'] = result.get('security_attribute', '')
        ws[f'H{row}'] = result.get('stride_model', '')
        ws[f'I{row}'] = result.get('threat_scenario', '')
        ws[f'J{row}'] = result.get('attack_path', '')
        ws[f'K{row}'] = result.get('wp29_mapping', '')
        
        # 威胁分析 - 攻击向量
        attack_vector = result.get('attack_vector', '本地')
        ws[f'L{row}'] = attack_vector
        ws[f'M{row}'] = f'=IF(L{row}="网络",0.85,IF(L{row}="邻居",0.62,IF(L{row}="本地",0.55,IF(L{row}="物理",0.2,0))))'
        
        # 攻击复杂度
        attack_complexity = result.get('attack_complexity', '低')
        ws[f'N{row}'] = attack_complexity
        ws[f'O{row}'] = f'=IF(N{row}="低",0.77,IF(N{row}="高",0.44,0))'
        
        # 权限要求
        privileges = result.get('privileges_required', '低')
        ws[f'P{row}'] = privileges
        ws[f'Q{row}'] = f'=IF(P{row}="无",0.85,IF(P{row}="低",0.62,IF(P{row}="高",0.27,0)))'
        
        # 用户交互
        user_interaction = result.get('user_interaction', '不需要')
        ws[f'R{row}'] = user_interaction
        ws[f'S{row}'] = f'=IF(R{row}="不需要",0.85,IF(R{row}="需要",0.62,0))'
        
        # 攻击可行性计算
        ws[f'T{row}'] = f'=8.22*M{row}*O{row}*Q{row}*S{row}'
        ws[f'U{row}'] = f'=IF(T{row}<=1.05,"很低",IF(T{row}<=1.99,"低",IF(T{row}<=2.99,"中",IF(T{row}<=3.99,"高","很高"))))'
        
        # 安全影响
        safety = result.get('safety_impact', '中等的')
        ws[f'V{row}'] = safety
        ws[f'W{row}'] = f'=IF(V{row}="可忽略不计的","没有受伤",IF(V{row}="中等的","轻伤和中等伤害",IF(V{row}="重大的","严重伤害(生存概率高)",IF(V{row}="严重的","危及生命(生存概率不确定)或致命伤害",""))))'
        ws[f'X{row}'] = f'=IF(V{row}="可忽略不计的",0,IF(V{row}="中等的",1,IF(V{row}="重大的",10,IF(V{row}="严重的",1000,0))))'
        
        # 经济影响
        financial = result.get('financial_impact', '中等的')
        ws[f'Y{row}'] = financial
        ws[f'Z{row}'] = f'=IF(Y{row}="可忽略不计的","财务损失不会产生任何影响",IF(Y{row}="中等的","财务损失会产生中等影响",IF(Y{row}="重大的","财务损失会产生重大影响",IF(Y{row}="严重的","财务损失会产生严重影响",""))))'
        ws[f'AA{row}'] = f'=IF(Y{row}="可忽略不计的",0,IF(Y{row}="中等的",1,IF(Y{row}="重大的",10,IF(Y{row}="严重的",1000,0))))'
        
        # 操作影响
        operational = result.get('operational_impact', '重大的')
        ws[f'AB{row}'] = operational
        ws[f'AC{row}'] = f'=IF(AB{row}="可忽略不计的","操作损坏不会导致车辆功能减少",IF(AB{row}="中等的","操作损坏会导致车辆功能中等减少",IF(AB{row}="重大的","操作损坏会导致车辆功能重大减少",IF(AB{row}="严重的","操作损坏会导致车辆功能丧失",""))))'
        ws[f'AD{row}'] = f'=IF(AB{row}="可忽略不计的",0,IF(AB{row}="中等的",1,IF(AB{row}="重大的",10,IF(AB{row}="严重的",1000,0))))'
        
        # 隐私影响
        privacy = result.get('privacy_impact', '可忽略不计的')
        ws[f'AE{row}'] = privacy
        ws[f'AF{row}'] = f'=IF(AE{row}="可忽略不计的","隐私危害不会产生任何影响",IF(AE{row}="中等的","隐私危害会产生中等影响",IF(AE{row}="重大的","隐私危害会产生重大影响",IF(AE{row}="严重的","隐私危害会产生严重影响",""))))'
        ws[f'AG{row}'] = f'=IF(AE{row}="可忽略不计的",0,IF(AE{row}="中等的",1,IF(AE{row}="重大的",10,IF(AE{row}="严重的",1000,0))))'
        
        # 影响等级计算
        ws[f'AH{row}'] = f'=SUM(X{row}+AA{row}+AD{row}+AG{row})'
        ws[f'AI{row}'] = f'=IF(AH{row}>=1000,"严重的",IF(AH{row}>=100,"重大的",IF(AH{row}>=10,"中等的",IF(AH{row}>=1,"可忽略不计的","无影响"))))'
        
        # 风险等级
        ws[f'AJ{row}'] = f'=IF(AND(AI{row}="无影响",U{row}="无"),"QM",IF(OR(AND(AI{row}="无影响",U{row}<>"无"),AND(AI{row}="可忽略不计的",OR(U{row}="很低",U{row}="低",U{row}="中")),AND(AI{row}="中等的",OR(U{row}="很低",U{row}="低")),AND(AI{row}="重大的",U{row}="很低")),"Low",IF(OR(AND(AI{row}="可忽略不计的",OR(U{row}="高",U{row}="很高")),AND(AI{row}="中等的",U{row}="中"),AND(AI{row}="重大的",U{row}="低"),AND(AI{row}="严重的",U{row}="很低")),"Medium",IF(OR(AND(AI{row}="中等的",OR(U{row}="高",U{row}="很高")),AND(AI{row}="重大的",U{row}="中"),AND(AI{row}="严重的",U{row}="低")),"High","Critical"))))'
        
        # 风险处置决策
        ws[f'AK{row}'] = f'=IF(OR(AJ{row}="QM",AJ{row}="Low"),"保留风险",IF(AJ{row}="Medium","降低风险","降低风险/规避风险/转移风险"))'
        
        # 安全目标和需求
        ws[f'AL{row}'] = f'=IF(AK{row}="保留风险","/",IF(OR(AK{row}="降低风险",AK{row}="降低风险/规避风险/转移风险"),"需要定义安全目标",""))'
        ws[f'AM{row}'] = result.get('security_requirement', '')
        
        # WP29控制映射
        ws[f'AN{row}'] = f'=IF(H{row}="T篡改","M10",IF(H{row}="D拒绝服务","M13",IF(H{row}="I信息泄露","M11",IF(H{row}="S欺骗","M23",IF(H{row}="R抵赖","M24",IF(H{row}="E权限提升","M16",""))))))'
        
        # 设置边框
        for col_idx in range(1, 41):
            col = get_column_letter(col_idx)
            ws[f'{col}{row}'].border = TARAStyles.THIN_BORDER
            ws[f'{col}{row}'].alignment = TARAStyles.CENTER_ALIGN if col_idx < 9 else TARAStyles.LEFT_ALIGN
        
        current_row += 1


# ==================== 主生成函数 ====================
def generate_tara_excel(
    output_path: str,
    cover_data: Dict[str, Any],
    definitions_data: Dict[str, Any],
    assets_data: Dict[str, Any],
    attack_trees_data: Dict[str, Any],
    tara_results_data: Dict[str, Any]
) -> str:
    """生成TARA分析报告Excel文件"""
    wb = Workbook()
    
    # 创建各个Sheet
    create_cover_sheet(wb, cover_data)
    create_definitions_sheet(wb, definitions_data)
    create_assets_sheet(wb, assets_data)
    create_attack_trees_sheet(wb, attack_trees_data)
    create_tara_results_sheet(wb, tara_results_data)
    
    # 保存文件
    wb.save(output_path)
    return output_path


def generate_tara_excel_from_json(
    output_path: str,
    json_data: Dict[str, Any]
) -> str:
    """从JSON数据生成TARA分析报告Excel文件"""
    return generate_tara_excel(
        output_path=output_path,
        cover_data=json_data.get('cover', {}),
        definitions_data=json_data.get('definitions', {}),
        assets_data=json_data.get('assets', {}),
        attack_trees_data=json_data.get('attack_trees', {}),
        tara_results_data=json_data.get('tara_results', {})
    )
